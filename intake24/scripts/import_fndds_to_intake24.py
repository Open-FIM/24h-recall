#!/usr/bin/env python3
"""
OpenFIM 24h Recall — FNDDS 2021-2023 to Intake24 Import Script
========================================================
Imports US foods from USDA FNDDS into Intake24 database
alongside existing UK foods (UK_current locale).
"""

import json
import openpyxl
import psycopg2
import uuid
import sys
import os

DB_CONFIG = {
    "host": os.environ.get("FOODS_DB_HOST", "127.0.0.1"),
    "port": int(os.environ.get("FOODS_DB_PORT", "5432")),
    "database": os.environ.get("FOODS_DB_NAME", "intake24_foods_dev"),
    "user": os.environ.get("FOODS_DB_USER", "postgres"),
    "password": os.environ.get("FOODS_DB_PASSWORD", ""),
}

FNDDS_DATA_DIR = os.environ.get("FNDDS_DATA_DIR", "intake24/data/fndds")

FOODS_FILE = os.path.join(FNDDS_DATA_DIR, "Foods_and_Beverages.xlsx")
NUTRIENTS_FILE = os.path.join(FNDDS_DATA_DIR, "FNDDS_Nutrient_Values.xlsx")
PORTIONS_FILE = os.path.join(FNDDS_DATA_DIR, "Portions_and_Weights.xlsx")

LOCALE_ID    = "en_US"
NUTRIENT_TABLE = "USDA"

NUTRIENT_MAP = {
    "Energy (kcal)":                          1,
    "Protein (g)":                           11,
    "Carbohydrate (g)":                      13,
    "Sugars, total\n(g)":                    22,
    "Fiber, total dietary (g)":              17,
    "Total Fat (g)":                         49,
    "Fatty acids, total saturated (g)":      50,
    "Fatty acids, total monounsaturated (g)":51,
    "Fatty acids, total polyunsaturated (g)":52,
    "Cholesterol (mg)":                      59,
    "Retinol (mcg)":                        114,
    "Vitamin A, RAE (mcg_RAE)":             283,
    "Carotene, alpha (mcg)":               116,
    "Carotene, beta (mcg)":                117,
    "Cryptoxanthin, beta (mcg)":           119,
    "Thiamin (mg)":                         123,
    "Riboflavin (mg)":                      124,
    "Niacin (mg)":                          125,
    "Vitamin B-6 (mg)":                     132,
    "Folic acid (mcg)":                     163,
    "Folate, food (mcg)":                   135,
    "Folate, DFE (mcg_DFE)":               162,
    "Folate, total (mcg)":                  134,
    "Vitamin B-12 (mcg)":                   133,
    "Vitamin C (mg)":                       129,
    "Vitamin D (D2 + D3) (mcg)":           122,
    "Vitamin E (alpha-tocopherol) (mg)":    155,
    "Vitamin K (phylloquinone) (mcg)":      177,
    "Calcium (mg)":                         140,
    "Phosphorus (mg)":                      142,
    "Magnesium (mg)":                       141,
    "Iron\n(mg)":                           143,
    "Zinc\n(mg)":                           147,
    "Copper (mg)":                          146,
    "Selenium (mcg)":                       152,
    "Potassium (mg)":                       139,
    "Sodium (mg)":                          138,
    "Caffeine (mg)":                        158,
    "Alcohol (g)":                           20,
    "Water\n(g)":                             8,
    "18:2\n(g)":                            100,
    "18:3\n(g)":                            103,
    "20:5 n-3\n(g)":                        111,
    "22:5 n-3\n(g)":                        112,
    "22:6 n-3\n(g)":                        113,
    "4:0\n(g)":                              70,
    "6:0\n(g)":                              71,
    "8:0\n(g)":                              72,
    "10:0\n(g)":                             73,
    "12:0\n(g)":                             74,
    "14:0\n(g)":                             75,
    "16:0\n(g)":                             77,
    "18:0\n(g)":                             79,
    "16:1\n(g)":                             86,
    "18:1\n(g)":                             89,
    "20:1\n(g)":                             92,
    "22:1\n(g)":                             95,
    "20:4\n(g)":                            110,
    "18:4\n(g)":                            105,
}

def load_foods(filepath):
    print(f"Loading foods...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    foods = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            foods.append({
                "food_code":    str(row[0]),
                "name":         str(row[1]),
                "alt_name":     str(row[2]) if row[2] else None,
                "category_num": row[3],
                "category_desc":str(row[4]) if row[4] else "Uncategorized"
            })
    wb.close()
    print(f"  {len(foods)} foods loaded")
    return foods

def load_nutrients(filepath):
    print(f"Loading nutrients...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[2]]
    nutrients = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            food_code = str(row[0])
            nutrients[food_code] = {}
            for i, header in enumerate(headers):
                if header in NUTRIENT_MAP and i < len(row) and row[i] is not None:
                    try:
                        nutrients[food_code][NUTRIENT_MAP[header]] = float(row[i])
                    except (ValueError, TypeError):
                        pass
    wb.close()
    print(f"  Nutrients loaded for {len(nutrients)} foods")
    return nutrients

def load_portions(filepath):
    print(f"Loading portions...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    portions = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[5] and row[6]:
            try:
                weight = float(row[6])
                if weight > 0:
                    food_code = str(row[0])
                    if food_code not in portions:
                        portions[food_code] = []
                    portions[food_code].append({
                        "description": str(row[5]),
                        "weight_g":    weight
                    })
            except (ValueError, TypeError):
                pass
    wb.close()
    print(f"  Portions loaded for {len(portions)} foods")
    return portions

def create_locale(conn):
    cur = conn.cursor()
    cur.execute("SELECT id FROM locales WHERE id = %s", (LOCALE_ID,))
    if cur.fetchone():
        print(f"  Locale {LOCALE_ID} already exists")
        return
    cur.execute("SELECT food_index_language_backend_id FROM locales WHERE id = 'UK_current'")
    row = cur.fetchone()
    lang_backend = row[0] if row else "en"
    cur.execute("""
        INSERT INTO locales (id, english_name, local_name, respondent_language_id,
            admin_language_id, country_flag_code, text_direction,
            food_index_language_backend_id, food_index_enabled)
        VALUES (%s, 'United States', 'United States', 'en', 'en', 'us', 'ltr', %s, true)
        ON CONFLICT (id) DO NOTHING
    """, (LOCALE_ID, lang_backend))
    conn.commit()
    print(f"  Locale {LOCALE_ID} created")

def create_categories(conn, foods):
    cur = conn.cursor()
    categories = {}
    for f in foods:
        if f["category_num"] and f["category_desc"]:
            categories[f["category_num"]] = f["category_desc"]
    created = 0
    for cat_num, cat_desc in categories.items():
        code = f"US{cat_num}"
        cur.execute("SELECT id FROM categories WHERE code = %s AND locale_id = %s",
                    (code, LOCALE_ID))
        if not cur.fetchone():
            cur.execute("""
                INSERT INTO categories (code, locale_id, name, english_name, simple_name, hidden, version)
                VALUES (%s, %s, %s, %s, %s, false, %s) ON CONFLICT DO NOTHING
            """, (code, LOCALE_ID, cat_desc, cat_desc, cat_desc.lower(), str(uuid.uuid4())))
            created += 1
    conn.commit()
    print(f"  {created} categories created ({len(categories)} total)")

def import_foods(conn, foods, nutrients, portions):
    cur = conn.cursor()
    total = len(foods)
    imported = 0
    skipped = 0

    for i, food in enumerate(foods):
        if i % 500 == 0:
            print(f"    {i}/{total} foods processed...")
            conn.commit()

        food_code = food["food_code"]
        code = f"US{food_code}"

        cur.execute("SELECT id FROM foods WHERE code = %s AND locale_id = %s",
                    (code, LOCALE_ID))
        if cur.fetchone():
            skipped += 1
            continue

        # Insert food
        alt_name_json = json.dumps({"en": food["alt_name"]}) if food.get("alt_name") else '{}'
        cur.execute("""
            INSERT INTO foods (code, locale_id, name, english_name, simple_name,
                version, alt_names, tags)
            VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, '[]')
            RETURNING id
        """, (code, LOCALE_ID, food["name"], food["name"],
              food["name"].lower(), str(uuid.uuid4()), alt_name_json))
        food_id = cur.fetchone()[0]

        # Link to category
        if food["category_num"]:
            cat_code = f"US{food['category_num']}"
            cur.execute("SELECT id FROM categories WHERE code = %s AND locale_id = %s",
                        (cat_code, LOCALE_ID))
            cat_row = cur.fetchone()
            if cat_row:
                cur.execute("""
                    INSERT INTO foods_categories (food_id, category_id)
                    VALUES (%s, %s) ON CONFLICT DO NOTHING
                """, (food_id, cat_row[0]))

        # Nutrient table record
        cur.execute("""
            INSERT INTO nutrient_table_records
                (nutrient_table_id, nutrient_table_record_id, name)
            VALUES (%s, %s, %s)
            ON CONFLICT (nutrient_table_id, nutrient_table_record_id) DO UPDATE
                SET name = EXCLUDED.name
            RETURNING id
        """, (NUTRIENT_TABLE, food_code, food["name"]))
        ntr_id = cur.fetchone()[0]

        # Link food to nutrient record
        cur.execute("""
            INSERT INTO foods_nutrients (food_id, nutrient_table_record_id)
            VALUES (%s, %s) ON CONFLICT DO NOTHING
        """, (food_id, ntr_id))

        # Nutrient values
        for nutrient_type_id, value in nutrients.get(food_code, {}).items():
            cur.execute("""
                INSERT INTO nutrient_table_record_nutrients
                    (nutrient_table_record_id, nutrient_type_id, units_per_100g)
                VALUES (%s, %s, %s) ON CONFLICT DO NOTHING
            """, (ntr_id, nutrient_type_id, value))

        # Portion size — pick best portion (prefer cup or oz)
        food_portions = portions.get(food_code, [])
        if not food_portions:
            food_portions = [{"description": "Standard serving (100g)", "weight_g": 100.0}]

        best = food_portions[0]
        for p in food_portions:
            desc = p["description"].lower()
            if "cup" in desc or "oz" in desc or "tbsp" in desc:
                best = p
                break

        cur.execute("""
            INSERT INTO food_portion_size_methods
                (food_id, method, description, conversion_factor, order_by,
                 parameters, pathways, default_weight)
            VALUES (%s, 'standard-unit', %s, 1.0, 0, %s, '["search"]', %s)
        """, (
            food_id,
            best["description"],
            json.dumps({"quantity": best["weight_g"]}),
            best["weight_g"]
        ))

        imported += 1

    conn.commit()
    print(f"  Import complete: {imported} imported, {skipped} already existed")

def main():
    print("=" * 60)
    print("OpenFIM 24h Recall — FNDDS 2021-2023 Import")
    print("=" * 60)

    for f in [FOODS_FILE, NUTRIENTS_FILE, PORTIONS_FILE]:
        if not os.path.exists(f):
            print(f"\nERROR: File not found: {f}")
            print("Please run:")
            print("  mkdir -p intake24/data/fndds")
            print("  # copy Excel files there and rename:")
            print("  #   Foods_and_Beverages.xlsx")
            print("  #   FNDDS_Nutrient_Values.xlsx")
            print("  #   Portions_and_Weights.xlsx")
            sys.exit(1)

    print("\nConnecting to database...")
    conn = psycopg2.connect(**DB_CONFIG)
    print("  Connected")

    print("\nStep 1: Loading FNDDS data from Excel files...")
    foods    = load_foods(FOODS_FILE)
    nutrients = load_nutrients(NUTRIENTS_FILE)
    portions = load_portions(PORTIONS_FILE)

    print(f"\nStep 2: Creating locale {LOCALE_ID}...")
    create_locale(conn)

    print(f"\nStep 3: Creating WWEIA categories...")
    create_categories(conn, foods)

    print(f"\nStep 4: Importing {len(foods)} foods with nutrients and portions...")
    import_foods(conn, foods, nutrients, portions)

    conn.close()
    print("\n" + "=" * 60)
    print("SUCCESS — FNDDS import complete!")
    print("=" * 60)
    print("\nNext steps:")
    print("1. Test: docker exec -it db psql -U postgres -d intake24_foods_dev -c \"SELECT COUNT(*) FROM foods WHERE locale_id = 'en_US';\"")
    print("2. Update survey locale in admin panel to include en_US")
    print("3. Restart the Intake24 API to rebuild food search index")

if __name__ == "__main__":
    main()